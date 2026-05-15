import sys
import re
import os
import shutil
import json
from datetime import datetime
import pandas as pd
import cv2
import numpy as np
from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, QLabel,
    QLineEdit, QPushButton, QTableWidget, QTableWidgetItem, QMessageBox,
    QFileDialog, QHeaderView, QDialog, QFormLayout, QGroupBox, QCheckBox,
    QStackedWidget, QComboBox, QDateEdit, QInputDialog, QFrame
)
from PyQt5.QtCore import Qt, QTimer
from PyQt5.QtGui import QPixmap

from Database import Database, IMAGE_FOLDER
from CameraDialog import CameraDialog
from ProductDetailPage import ProductDetailPage
from AllProductsPage import AllProductsPage
from StatisticsPage import StatisticsPage
from BaiduOCRDialog import BaiduOCRDialog
from AdminWindow_def import AdminUtils

DEFAULT_LOCATION = "湖北省 武汉市"


# ---------- 数据库管理对话框 ----------
class DatabaseManagerDialog(QDialog):
    def __init__(self, parent, db_manager):
        super().__init__(parent)
        self.setWindowTitle("数据库管理")
        self.setMinimumWidth(400)
        self.db_manager = db_manager
        self.init_ui()
        self.load_current_config()

    def init_ui(self):
        layout = QVBoxLayout()

        switch_layout = QHBoxLayout()
        self.mysql_switch = QCheckBox("启用 MySQL 服务器")
        self.mysql_switch.setStyleSheet("""
            QCheckBox::indicator {
                width: 40px;
                height: 20px;
                border-radius: 10px;
                background-color: #cccccc;
            }
            QCheckBox::indicator:checked {
                background-color: #4CAF50;
            }
        """)
        self.mysql_switch.stateChanged.connect(self.on_switch_changed)
        switch_layout.addWidget(self.mysql_switch)
        switch_layout.addStretch()
        layout.addLayout(switch_layout)

        self.config_widget = QWidget()
        config_layout = QFormLayout(self.config_widget)
        self.host_edit = QLineEdit("10.99.0.107")
        self.user_edit = QLineEdit("jms")
        self.password_edit = QLineEdit()
        self.password_edit.setEchoMode(QLineEdit.Password)
        self.db_edit = QLineEdit("jewelry_db")
        config_layout.addRow("服务器 IP:", self.host_edit)
        config_layout.addRow("用户名:", self.user_edit)
        config_layout.addRow("密码:", self.password_edit)
        config_layout.addRow("数据库名:", self.db_edit)

        btn_row = QHBoxLayout()
        self.save_config_btn = QPushButton("💾 保存配置")
        self.save_config_btn.clicked.connect(self.save_config)
        self.test_conn_btn = QPushButton("🔍 测试连接")
        self.test_conn_btn.clicked.connect(self.test_connection)
        btn_row.addWidget(self.save_config_btn)
        btn_row.addWidget(self.test_conn_btn)
        config_layout.addRow(btn_row)
        layout.addWidget(self.config_widget)

        line = QFrame()
        line.setFrameShape(QFrame.HLine)
        line.setFrameShadow(QFrame.Sunken)
        layout.addWidget(line)

        self.merge_btn = QPushButton("🗄️ 合并数据库")
        self.merge_btn.clicked.connect(self.merge_databases)
        self.import_btn = QPushButton("⬆️ 导入旧版数据")
        self.import_btn.clicked.connect(self.import_old)
        layout.addWidget(self.merge_btn)
        layout.addWidget(self.import_btn)

        close_btn = QPushButton("关闭")
        close_btn.clicked.connect(self.close)
        layout.addWidget(close_btn, alignment=Qt.AlignRight)

        self.setLayout(layout)

    def load_current_config(self):
        path = os.path.join(os.path.dirname(__file__), "mysql_config.json")
        if os.path.exists(path):
            try:
                with open(path, 'r', encoding='utf-8') as f:
                    cfg = json.load(f)
                self.host_edit.setText(cfg.get("host", ""))
                self.user_edit.setText(cfg.get("user", ""))
                self.password_edit.setText(cfg.get("password", ""))
                self.db_edit.setText(cfg.get("database", ""))
                self.mysql_switch.setChecked(True)
            except:
                self.mysql_switch.setChecked(False)
        else:
            self.mysql_switch.setChecked(False)
        self.on_switch_changed()

    def on_switch_changed(self):
        enabled = self.mysql_switch.isChecked()
        self.config_widget.setVisible(enabled)
        if not enabled:
            self._delete_config_file()
        else:
            pass

    def _delete_config_file(self):
        path = os.path.join(os.path.dirname(__file__), "mysql_config.json")
        if os.path.exists(path):
            try:
                os.remove(path)
            except Exception as e:
                QMessageBox.warning(self, "提示", f"删除配置文件失败: {e}")

    def save_config(self):
        config = {
            "host": self.host_edit.text().strip(),
            "user": self.user_edit.text().strip(),
            "password": self.password_edit.text(),
            "database": self.db_edit.text().strip()
        }
        path = os.path.join(os.path.dirname(__file__), "mysql_config.json")
        with open(path, 'w', encoding='utf-8') as f:
            json.dump(config, f, ensure_ascii=False, indent=2)
        QMessageBox.information(self, "成功", "配置已保存，下次启动程序将使用该连接。")

    def test_connection(self):
        try:
            import pymysql
            conn = pymysql.connect(
                host=self.host_edit.text().strip(),
                user=self.user_edit.text().strip(),
                password=self.password_edit.text(),
                database=self.db_edit.text().strip(),
                charset='utf8mb4',
                connect_timeout=5
            )
            conn.close()
            QMessageBox.information(self, "成功", "MySQL 连接成功！")
        except Exception as e:
            QMessageBox.critical(self, "连接失败", str(e))

    def merge_databases(self):
        file_paths, _ = QFileDialog.getOpenFileNames(
            self, "选择要合并的数据库文件", "",
            "SQLite 数据库 (*.db)"
        )
        if not file_paths:
            return
        reply = QMessageBox.question(
            self, "确认合并",
            f"您选择了 {len(file_paths)} 个文件，是否开始合并？\n已存在的 ID 将自动跳过。",
            QMessageBox.Yes | QMessageBox.No
        )
        if reply != QMessageBox.Yes:
            return
        try:
            inserted, skipped, errors = AdminUtils.merge_sqlite_files_to_current_db(
                self.db_manager.db, file_paths
            )
            msg = f"合并完成！\n成功插入: {inserted} 条\n跳过重复: {skipped} 条"
            if errors:
                msg += f"\n错误: {len(errors)} 条（详见控制台）"
                for err in errors[:5]:
                    print(err)
            QMessageBox.information(self, "合并结果", msg)
            self._refresh_pages()
        except Exception as e:
            QMessageBox.critical(self, "合并失败", str(e))

    def import_old(self):
        reply = QMessageBox.information(
            self, "导入旧版数据库",
            "此功能用于将旧版本数据库合并到当前数据库。\n"
            "系统会自动处理字段差异，重复 ID 将跳过。\n\n"
            "请选择要导入的 .db 文件。",
            QMessageBox.Ok | QMessageBox.Cancel
        )
        if reply != QMessageBox.Ok:
            return
        file_paths, _ = QFileDialog.getOpenFileNames(
            self, "选择旧版数据库文件", "",
            "SQLite 数据库 (*.db)"
        )
        if not file_paths:
            return
        try:
            inserted, skipped, errors = AdminUtils.merge_sqlite_files_to_current_db(
                self.db_manager.db, file_paths
            )
            msg = f"导入完成！\n成功: {inserted} 条\n跳过: {skipped} 条"
            if errors:
                msg += f"\n错误: {len(errors)} 条（详见控制台）"
                for err in errors[:5]:
                    print(err)
            QMessageBox.information(self, "导入结果", msg)
            self._refresh_pages()
        except Exception as e:
            QMessageBox.critical(self, "导入失败", str(e))

    def _refresh_pages(self):
        main = self.db_manager
        if main.stacked_widget.currentIndex() == 1:
            main.all_products_page.load_data()
        elif main.stacked_widget.currentIndex() == 2:
            main.detail_page.load_product_data()


# ================== 主窗口 ==================
class AdminWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("首饰管理系统-管理员-v1.2.0")
        self.setGeometry(100, 100, 1000, 700)

        self.setStyleSheet("""
            QMainWindow { background-color: #F5FBFB; }
            QWidget#navWidget {
                background-color: #DDDEDE;
                border-right: 2px solid #02D3D3;
            }
            QPushButton {
                background-color: #F5FBFB;
                color: #2c3e50;
                border: 1px solid #02D3D3;
                padding: 10px 16px;
                text-align: left;
                font-size: 14px;
                font-weight: 500;
                border-radius: 6px;
                margin: 3px 10px;
            }
            QPushButton:hover { background-color: #E0F7FA; border-color: #00BCD4; }
            QPushButton:pressed { background-color: #B2EBF2; }
            QPushButton:checked {
                background-color: #02D3D3;
                color: #FFFFFF;
                border-color: #00ACC1;
                font-weight: bold;
            }
            QPushButton:disabled {
                color: #7f8c8d;
                background-color: #F5FBFB;
                border-color: #DDDEDE;
            }
            QWidget#rightContainer { background-color: #FFFFFF; }
            QLabel#timeLabel {
                background-color: #DDDEDE;
                color: #2c3e50;
                font-size: 20px;
                padding: 6px 15px;
                border-bottom: 2px solid #02D3D3;
            }
            QGroupBox {
                font-size: 14px;
                font-weight: bold;
                color: #2c3e50;
                border: 1px solid #DDDEDE;
                border-radius: 8px;
                margin-top: 10px;
                padding-top: 15px;
                background-color: #FFFFFF;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 15px;
                padding: 0 5px;
                color: #02D3D3;
            }
            QLineEdit, QComboBox, QDateEdit {
                border: 1px solid #DDDEDE;
                border-radius: 5px;
                padding: 6px 10px;
                background-color: #F5FBFB;
                font-size: 16px;
                min-height: 30px;
            }
            QLineEdit:focus, QComboBox:focus, QDateEdit:focus { border-color: #02D3D3; }
            QComboBox::drop-down {
                subcontrol-origin: padding;
                subcontrol-position: top right;
                width: 20px;
                border-left: 1px solid #DDDEDE;
            }
            QPushButton#pageBtn {
                background-color: #02D3D3;
                color: #FFFFFF;
                border: 1px solid #00ACC1;
                border-radius: 6px;
                padding: 8px 16px;
                font-size: 14px;
                font-weight: bold;
            }
            QPushButton#pageBtn:hover { background-color: #00BCD4; border-color: #0097A7; }
            QPushButton#pageBtn:pressed { background-color: #0097A7; }
            QTableWidget {
                background-color: #FFFFFF;
                alternate-background-color: #F5FBFB;
                gridline-color: #DDDEDE;
                selection-background-color: #B2EBF2;
                selection-color: #2c3e50;
                border: 1px solid #DDDEDE;
                border-radius: 4px;
            }
            QHeaderView::section {
                background-color: #DDDEDE;
                padding: 6px;
                border: none;
                border-bottom: 1px solid #02D3D3;
                font-weight: bold;
                color: #2c3e50;
            }
        """)

        if not os.path.exists(IMAGE_FOLDER):
            os.makedirs(IMAGE_FOLDER)

        self.db = Database()
        self.current_image_path = None

        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QHBoxLayout()
        central_widget.setLayout(main_layout)

        # 左侧导航栏
        nav_widget = QWidget()
        nav_widget.setObjectName("navWidget")
        nav_widget.setFixedWidth(250)
        nav_layout = QVBoxLayout()
        nav_widget.setLayout(nav_layout)

        self.btn_add_page = QPushButton("➕ 首饰录入")
        self.btn_add_page.setCheckable(True)
        self.btn_add_page.clicked.connect(lambda: self.switch_page(0))

        self.btn_all_products = QPushButton("📋 所有首饰列表")
        self.btn_all_products.setCheckable(True)
        self.btn_all_products.clicked.connect(lambda: self.switch_page(1))

        self.btn_detail_page = QPushButton("📝 查询/编辑/删除")
        self.btn_detail_page.setCheckable(True)
        self.btn_detail_page.clicked.connect(lambda: self.switch_page(2))

        self.btn_stats_page = QPushButton("📊 统计报表")
        self.btn_stats_page.setCheckable(True)
        self.btn_stats_page.clicked.connect(lambda: self.switch_page(3))

        self.btn_export = QPushButton("📊 导出 Excel")
        self.btn_export.clicked.connect(self.export_excel)

        self.btn_baidu_ocr = QPushButton("✍️ 手写识别(百度)")
        self.btn_baidu_ocr.clicked.connect(self.open_baidu_ocr)

        self.btn_match_image = QPushButton("📷 拍照匹配首饰")
        self.btn_match_image.clicked.connect(self.open_image_matcher)

        # 数据库管理按钮（整合原三个按钮）
        self.btn_db_manager = QPushButton("🗄️ 数据库管理")
        self.btn_db_manager.clicked.connect(self.open_database_manager)

        nav_layout.addWidget(self.btn_add_page)
        nav_layout.addWidget(self.btn_baidu_ocr)
        nav_layout.addWidget(self.btn_match_image)
        nav_layout.addWidget(self.btn_all_products)
        nav_layout.addWidget(self.btn_detail_page)
        nav_layout.addWidget(self.btn_stats_page)
        nav_layout.addWidget(self.btn_export)
        nav_layout.addWidget(self.btn_db_manager)
        nav_layout.addStretch()
        main_layout.addWidget(nav_widget)

        # 右侧区域
        right_container = QWidget()
        right_container.setObjectName("rightContainer")
        right_layout = QVBoxLayout(right_container)
        right_layout.setContentsMargins(0, 0, 0, 0)
        right_layout.setSpacing(5)

        # 将 time_label 和 db_type_label 放在同一个水平布局中
        self.time_layout = QHBoxLayout()
        self.time_layout.setContentsMargins(0, 0, 0, 0)

        # 数据库类型标签（左侧）
        self.db_type_label = QLabel()
        self.db_type_label.setStyleSheet("font-size: 12pt; padding: 5px 10px; color: #2c3e50; font-weight: bold;")
        self.time_layout.addWidget(self.db_type_label)
        self.time_layout.addStretch()

        # 时间标签（右侧）
        self.time_label = QLabel()
        self.time_label.setObjectName("timeLabel")
        self.time_label.setAlignment(Qt.AlignRight)
        self.time_layout.addWidget(self.time_label)

        right_layout.addLayout(self.time_layout)  # 注意这里改为 addLayout，之前是 addWidget

        self.stacked_widget = QStackedWidget()
        right_layout.addWidget(self.stacked_widget)

        main_layout.addWidget(right_container, 1)

        self.add_page = self.create_add_page()
        self.stacked_widget.addWidget(self.add_page)

        self.all_products_page = AllProductsPage(self.db, self)
        self.stacked_widget.addWidget(self.all_products_page)

        self.detail_page = ProductDetailPage(self.db, self)
        self.stacked_widget.addWidget(self.detail_page)

        self.stats_page = StatisticsPage(self.db, self)
        self.stacked_widget.addWidget(self.stats_page)

        self.btn_add_page.setChecked(True)

        self.timer = QTimer(self)
        self.timer.timeout.connect(self.update_time)
        self.timer.start(1000)
        self.update_time()

        self.refresh_all_lists()

        # 如果用户配置了 MySQL 但连接失败，弹出温馨提示
        if self.db.mysql_failed:
            QMessageBox.information(
                self, "数据库提示",
                "MySQL 服务器连接失败，已自动切换到本地 SQLite 数据库。\n"
                "如需使用 MySQL，请检查服务器状态或重新配置。"
            )

    # ---------- 打开数据库管理对话框 ----------
    def open_database_manager(self):
        dialog = DatabaseManagerDialog(self, self)
        dialog.exec_()

        # 重新创建数据库连接（自动读取最新的 mysql_config.json）
        old_db = self.db
        self.db = Database()

        # 更新所有子页面的数据库引用
        self.all_products_page.db = self.db
        self.detail_page.db = self.db
        self.stats_page.db = self.db

        # 关闭旧连接
        if old_db is not None:
            try:
                old_db.close()
            except:
                pass

        # 刷新标题和所有页面数据
        self.update_time()  # 立刻更新数据库类型标签
        self.all_products_page.load_data()
        self.detail_page.load_product_data()
        self.stats_page.load_all_time_stats()

    # ---------- 草稿相关（保持不变） ----------
    def _draft_path(self):
        return os.path.join(os.path.dirname(__file__), "draft.json")

    def _draft_exists(self):
        return os.path.exists(self._draft_path())

    def has_unsaved_changes(self):
        if (self.name_input.text().strip() or
            self.sell_input.text().strip() or
            self.cost_input.text().strip() or
            self.remark_input.text().strip()):
            return True
        if hasattr(self, 'current_image_path') and self.current_image_path:
            return True
        return False

    def save_draft(self):
        draft = {
            'name': self.name_input.text(),
            'sell': self.sell_input.text(),
            'cost': self.cost_input.text(),
            'remark': self.remark_input.text(),
            'brand': self.brand_combo.currentText(),
            'category': self.category_combo.currentText(),
            'image_path': self.current_image_path if self.current_image_path else ''
        }
        try:
            with open(self._draft_path(), 'w', encoding='utf-8') as f:
                json.dump(draft, f, ensure_ascii=False, indent=2)
            print("草稿保存成功")
        except Exception as e:
            print(f"草稿保存失败: {e}")

    def load_draft(self):
        path = self._draft_path()
        if not os.path.exists(path):
            return False
        try:
            with open(path, 'r', encoding='utf-8') as f:
                draft = json.load(f)
            self.name_input.setText(draft.get('name', ''))
            self.sell_input.setText(draft.get('sell', ''))
            self.cost_input.setText(draft.get('cost', ''))
            self.remark_input.setText(draft.get('remark', ''))
            brand = draft.get('brand', '')
            if brand:
                if self.brand_combo.findText(brand) == -1:
                    self.brand_combo.addItem(brand)
                self.brand_combo.setCurrentText(brand)
            else:
                self.brand_combo.setCurrentIndex(0)
            category = draft.get('category', '')
            if category:
                if self.category_combo.findText(category) == -1:
                    self.category_combo.addItem(category)
                self.category_combo.setCurrentText(category)
            else:
                self.category_combo.setCurrentIndex(0)
            img_path = draft.get('image_path', '')
            if img_path and os.path.exists(img_path):
                self.current_image_path = img_path
                self.file_img_btn.setText(f"📂 已选择: {os.path.basename(img_path)}")
            else:
                self.current_image_path = None
                self.file_img_btn.setText("📂 选择文件...")
            print("草稿恢复成功")
            return True
        except Exception as e:
            print(f"加载草稿失败: {e}")
            return False

    def clear_draft_file(self):
        path = self._draft_path()
        if os.path.exists(path):
            try:
                os.remove(path)
            except Exception as e:
                print(f"删除草稿文件失败: {e}")

    def clear_draft_fields(self):
        self.name_input.clear()
        self.sell_input.clear()
        self.cost_input.clear()
        self.remark_input.clear()
        self.brand_combo.setCurrentIndex(0)
        self.category_combo.setCurrentIndex(0)
        self.file_img_btn.setText("📂 选择文件...")
        self.current_image_path = None
        self.clear_draft_file()

    def check_draft_on_leave(self):
        if self.has_unsaved_changes():
            reply = QMessageBox.question(
                self, "未保存的更改",
                "当前录入内容尚未保存，是否保存为草稿？\n（草稿可在下次打开时恢复）",
                QMessageBox.Yes | QMessageBox.No | QMessageBox.Cancel
            )
            if reply == QMessageBox.Yes:
                self.save_draft()
                return True
            elif reply == QMessageBox.No:
                self.clear_draft_file()
                return True
            else:
                return False
        return True

    def closeEvent(self, event):
        if self.stacked_widget.currentIndex() == 0:
            if not self.check_draft_on_leave():
                event.ignore()
                return
        event.accept()

    def switch_page(self, index):
        if self.stacked_widget.currentIndex() == 0 and index != 0:
            if not self.check_draft_on_leave():
                return
        self.stacked_widget.setCurrentIndex(index)
        self.btn_add_page.setChecked(index == 0)
        self.btn_all_products.setChecked(index == 1)
        self.btn_detail_page.setChecked(index == 2)
        self.btn_stats_page.setChecked(index == 3)
        if index == 0:
            self.refresh_all_lists()
            if self._draft_exists():
                reply = QMessageBox.question(
                    self, "恢复草稿",
                    "检测到未完成的草稿，是否恢复？",
                    QMessageBox.Yes | QMessageBox.No
                )
                if reply == QMessageBox.Yes:
                    self.load_draft()
                    return
                else:
                    self.clear_draft_fields()
        elif index == 1:
            self.all_products_page.load_data()
        elif index == 3:
            self.stats_page.load_all_time_stats()

    # ---------- 业务方法（保持不变） ----------
    def update_time(self):
        current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        self.time_label.setText(f"⏰ 当前时间：{current_time}")
        # 根据实际数据库连接状态更新标签
        if self.db.use_mysql:
            self.db_type_label.setText("🗄️ 外部数据库 (MySQL)")
        else:
            self.db_type_label.setText("📂 本地数据库 (SQLite)")

    def create_add_page(self):
        page = QWidget()
        layout = QVBoxLayout()
        input_group = QGroupBox("新首饰录入")
        form_layout = QFormLayout()

        self.name_input = QLineEdit()
        self.sell_input = QLineEdit()
        self.cost_input = QLineEdit()
        self.remark_input = QLineEdit()

        style = "QLineEdit { min-height: 30px; font-size: 12pt; }"
        for inp in [self.name_input, self.sell_input, self.cost_input, self.remark_input]:
            inp.setStyleSheet(style)

        self.category_combo = QComboBox()
        self.category_combo.setEditable(True)
        self.category_combo.setPlaceholderText("选择或输入新品类")
        add_cat_btn = QPushButton("➕")
        add_cat_btn.setFixedSize(25, 25)
        add_cat_btn.clicked.connect(self.add_new_category)
        cat_layout = QHBoxLayout()
        cat_layout.addWidget(self.category_combo)
        cat_layout.addWidget(add_cat_btn)
        form_layout.addRow("品类:", cat_layout)

        self.brand_combo = QComboBox()
        self.brand_combo.setEditable(True)
        self.brand_combo.setPlaceholderText("选择或输入新品牌")
        add_brand_btn = QPushButton("➕")
        add_brand_btn.setFixedSize(25, 25)
        add_brand_btn.clicked.connect(self.add_new_brand)
        brand_layout = QHBoxLayout()
        brand_layout.addWidget(self.brand_combo)
        brand_layout.addWidget(add_brand_btn)
        form_layout.addRow("品牌:", brand_layout)

        form_layout.addRow("首饰名称:", self.name_input)
        form_layout.addRow("售价:", self.sell_input)
        form_layout.addRow("进价:", self.cost_input)
        form_layout.addRow("备注:", self.remark_input)

        img_btn_layout = QHBoxLayout()
        self.file_img_btn = QPushButton("📂 选择文件...")
        self.file_img_btn.setStyleSheet("min-height: 30px;")
        self.file_img_btn.clicked.connect(self.upload_image)
        img_btn_layout.addWidget(self.file_img_btn)

        self.camera_btn = QPushButton("📷 拍照录入")
        self.camera_btn.setStyleSheet("min-height: 30px;")
        self.camera_btn.clicked.connect(self.open_camera)
        img_btn_layout.addWidget(self.camera_btn)

        form_layout.addRow("图片上传:", img_btn_layout)

        self.add_btn = QPushButton("确认录入")
        self.add_btn.setObjectName("pageBtn")
        self.add_btn.clicked.connect(self.add_product)

        form_layout.addRow(self.add_btn)
        input_group.setLayout(form_layout)
        layout.addWidget(input_group)

        self.refresh_all_lists()
        page.setLayout(layout)
        return page

    def refresh_all_lists(self):
        brands = self.db.get_all_brands()
        current_brand = self.brand_combo.currentText()
        self.brand_combo.blockSignals(True)
        self.brand_combo.clear()
        self.brand_combo.addItem("")
        self.brand_combo.addItems([b for b in brands if not (b and ('-' in b or ':' in b) and len(b) > 10)])
        if current_brand and self.brand_combo.findText(current_brand) != -1:
            self.brand_combo.setCurrentText(current_brand)
        else:
            self.brand_combo.setCurrentIndex(0)
        self.brand_combo.blockSignals(False)

        categories = self.db.get_all_categories()
        current_cat = self.category_combo.currentText()
        self.category_combo.blockSignals(True)
        self.category_combo.clear()
        self.category_combo.addItem("")
        self.category_combo.addItems(categories)
        if current_cat and self.category_combo.findText(current_cat) != -1:
            self.category_combo.setCurrentText(current_cat)
        else:
            self.category_combo.setCurrentIndex(0)
        self.category_combo.blockSignals(False)

    def add_new_brand(self):
        brand, ok = QInputDialog.getText(self, "添加新品牌", "请输入新品牌名称：")
        if ok and brand.strip():
            self.brand_combo.addItem(brand.strip())
            self.brand_combo.setCurrentText(brand.strip())
            QMessageBox.information(self, "提示", f"品牌「{brand.strip()}」已添加到当前输入框。")

    def add_new_category(self):
        category, ok = QInputDialog.getText(self, "添加新品类", "请输入新品类名称：")
        if ok and category.strip():
            self.category_combo.addItem(category.strip())
            self.category_combo.setCurrentText(category.strip())
            QMessageBox.information(self, "提示", f"品类「{category.strip()}」已添加到当前输入框。")

    def upload_image(self):
        file_path, _ = QFileDialog.getOpenFileName(self, "选择图片", "", "Images (*.png *.jpg *.jpeg)")
        if file_path:
            self.current_image_path = file_path
            self.file_img_btn.setText(f"📂 已选择: {os.path.basename(file_path)}")

    def open_camera(self):
        if not os.path.exists(IMAGE_FOLDER):
            os.makedirs(IMAGE_FOLDER)
        dialog = CameraDialog(self)
        dialog.exec_()

    def set_image_from_camera(self, path):
        self.current_image_path = path
        self.file_img_btn.setText(f"📷 已拍照: {os.path.basename(path)}")

    def add_product(self):
        name = self.name_input.text().strip()
        sell_text = self.sell_input.text().strip()
        cost_text = self.cost_input.text().strip()
        remark = self.remark_input.text().strip()
        brand = self.brand_combo.currentText().strip()
        category = self.category_combo.currentText().strip()
        if not name or not sell_text or not cost_text:
            QMessageBox.warning(self, "警告", "请填写完整首饰信息")
            return
        try:
            sell_price = float(sell_text)
            cost_price = float(cost_text)
        except ValueError:
            QMessageBox.warning(self, "警告", "售价和进价必须是数字")
            return
        product_id = AdminUtils.generate_id()
        final_image_path = ""
        if self.current_image_path:
            product_dir = os.path.join(IMAGE_FOLDER, product_id)
            os.makedirs(product_dir, exist_ok=True)
            original_name = os.path.basename(self.current_image_path)
            safe_name = re.sub(r'[\\/*?:"<>|]', '', original_name)
            target_path = os.path.join(product_dir, safe_name)
            counter = 1
            while os.path.exists(target_path):
                name_part, ext_part = os.path.splitext(safe_name)
                target_path = os.path.join(product_dir, f"{name_part}_{counter}{ext_part}")
                counter += 1
            shutil.copy(self.current_image_path, target_path)
            final_image_path = f"{product_id}/{os.path.basename(target_path)}"
            ext = os.path.splitext(self.current_image_path)[1]
            safe_brand = re.sub(r'[\\/*?:"<>|]', '', brand) if brand else "无品牌"
            safe_name_brand = re.sub(r'[\\/*?:"<>|]', '', name) if name else "无名称"
            brand_name = f"{safe_brand}-{safe_name_brand}"
            brand_target = os.path.join(IMAGE_FOLDER, f"{brand_name}{ext}")
            counter = 1
            while os.path.exists(brand_target):
                brand_target = os.path.join(IMAGE_FOLDER, f"{brand_name}_{counter}{ext}")
                counter += 1
            shutil.copy(self.current_image_path, brand_target)

        create_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        data = (
            product_id, name, final_image_path, sell_price, cost_price, remark,
            DEFAULT_LOCATION, create_time, "", "", 0, "", brand, category
        )
        if self.db.insert_product(data):
            QMessageBox.information(self, "成功", f"首饰录入成功！\n编号: {product_id}\n地点: {DEFAULT_LOCATION}")
            self.clear_draft_fields()
            if final_image_path:
                full = os.path.join(IMAGE_FOLDER, final_image_path)
                if os.path.exists(full):
                    self.db.update_image_hash(product_id, AdminUtils.compute_image_hash_from_path(full))
            if self.stacked_widget.currentIndex() == 1:
                self.all_products_page.load_data()
            if self.stacked_widget.currentIndex() == 3:
                self.stats_page.load_all_time_stats()
        else:
            QMessageBox.critical(self, "错误", "录入失败，可能是ID重复")

    def export_excel(self):
        try:
            import pandas as pd
            from openpyxl import load_workbook
            from openpyxl.utils import get_column_letter
            from openpyxl.styles import Alignment
        except ImportError:
            QMessageBox.critical(self, "错误", "缺少必要的库，请安装 pandas 和 openpyxl")
            return
        products = self.db.get_all_products()
        if not products:
            QMessageBox.information(self, "提示", "没有数据可导出")
            return
        cleaned = []
        for p in products:
            lst = list(p)
            if len(lst) > 14: lst = lst[:14]
            elif len(lst) < 14: lst.extend([""] * (14 - len(lst)))
            cleaned.append([str(x) if x is not None else "" for x in lst])
        df = pd.DataFrame(cleaned, columns=[
            "编号", "名称", "图片路径", "售价", "进价", "备注", "地点", "录入时间",
            "平台", "描述", "是否售出", "售出方式", "品牌", "品类"
        ])
        df = df[["名称", "品类", "品牌", "编号", "售价", "进价", "备注", "录入时间", "地点"]]
        filename = f"商品库存表_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        abs_path = os.path.abspath(filename)
        df.to_excel(abs_path, index=False, engine='openpyxl')
        wb = load_workbook(abs_path)
        ws = wb.active
        for row in ws.iter_rows():
            ws.row_dimensions[row[0].row].height = 10 * (72 / 25.4)
        for col in range(1, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(col)].width = 20
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')
        wb.save(abs_path)
        QMessageBox.information(self, "成功", f"Excel导出成功！\n文件保存在:\n{abs_path}")

    def open_baidu_ocr(self):
        try:
            dialog = BaiduOCRDialog(self.db, self, api_key="Za5vLSHG5se97PlKJctfMWUZ",
                                   secret_key="ciD2mt1pa4j7MrA3OqRUdo2viVMwbRVN")
            dialog.exec_()
        except Exception as e:
            QMessageBox.critical(self, "功能暂时不可用", f"匹配模块加载失败，已禁用。\n错误: {e}")
            self.btn_match_image.setEnabled(False)
            self.btn_match_image.setText("📷 匹配(已禁用)")

    def open_image_matcher(self):
        try:
            from ImageMatcherDialog import ImageMatcherDialog
            dialog = ImageMatcherDialog(self.db, self)
            dialog.exec_()
        except Exception as e:
            QMessageBox.critical(self, "功能暂时不可用", f"匹配模块加载失败，已禁用。\n错误: {e}")
            self.btn_match_image.setEnabled(False)
            self.btn_match_image.setText("📷 匹配(已禁用)")